import sys
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side


rows = ['Emails Sent', 'Emails Delivered', 'Undeliverable', 'Survey Responses',
        'Total Click Throughs', 'Unique Click Throughs', 'Unique Emails Opened', 'Unsubscribes',
        'FTAF-Forwarders', 'FTAF-Recipients', 'FTAF-Subscribers', 'Open Rate',
        'Deliverability Rate', 'Bounce Rate', 'Unsubscribe Rate', 'Unique Click Through Rate',
        'Unique Complaints', 'Cumulative Complaints']
sum_rows = ['Emails Sent', 'Emails Delivered', 'Undeliverable', 'Survey Responses',
            'Total Click Throughs', 'Unique Click Throughs', 'Unique Emails Opened', 'Unsubscribes',
            'FTAF-Forwarders', 'FTAF-Recipients', 'FTAF-Subscribers', 'Unique Complaints',
            'Cumulative Complaints']
avg_rows = ['Open Rate', 'Deliverability Rate', 'Bounce Rate', 'Unsubscribe Rate',
            'Unique Click Through Rate']


def read_excel(path_to_file):
    xls = pd.ExcelFile(path_to_file, engine='openpyxl')
    df = xls.parse(xls.sheet_names[0], skiprows=7,
                   index_col=None, na_values=['NA'])

    xls.close()

    return df


def cal_report(df):
    df_sum_temp = df.groupby(['Name'])[sum_rows].agg('sum')
    df_avg_temp = df.groupby(['Name'])[avg_rows].agg('mean')
    for row in avg_rows:
        df_avg_temp[row] = pd.Series(["{0:.2f}%".format(
            val * 100) for val in df_avg_temp[row]], index=df_avg_temp.index)

    return df_sum_temp, df_avg_temp


def rearrange_df(df_sum_temp, df_avg_temp):
    df = df_sum_temp.merge(df_avg_temp, left_on="Name", right_on="Name")
    df = df[rows]
    return df


def check_result_sheet(path_to_file):
    wb = openpyxl.load_workbook(path_to_file)
    if 'Result' in wb.sheetnames:
        std=wb['Result']
        wb.remove(std)
    wb.save(path_to_file)
    wb.close()


def write_to_excel(df, path_to_file):
    check_result_sheet(path_to_file)
    writer = pd.ExcelWriter(path_to_file, mode='a',engine='openpyxl')
    df.to_excel(writer, sheet_name='Result')
    writer.save()
    writer.close()


def format_excel(path_to_file, df):
    wb = load_workbook(path_to_file)
    num_col = len(df.columns)

    # Define style for Header
    wb_style_header = NamedStyle('wb_style_header')
    font = Font(name='Arial', size=8, bold=True, color='FFFFFF')
    wb_style_header.font = font
    bd = Side(style='thin', color="000000")
    wb_style_header.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb_style_header.fill = PatternFill(fgColor="455560", fill_type="solid")
    wb_style_header.alignment = Alignment(horizontal='center')
    wb_style_header.alignment.vertical = "top"

    # Define style for Cells
    wb_style_cells = NamedStyle('wb_style_cells')
    font = Font(name='Arial', size=8)
    wb_style_cells.font = font
    wb_style_cells.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    
    # Delete styles if exists
    styles = ['wb_style_header', 'wb_style_cells']
    for style in styles:
        if style in wb.named_styles:
            del wb._named_styles[ wb.style_names.index(style)]

    ws = wb[wb.sheetnames[1]]
    ws.auto_filter.ref = ws.dimensions

    #Apply styles
    for i in range(num_col+1):
        ws[f'{chr(ord("A") + i)}1'].style = wb_style_header

    for i in range(num_col+1):
        for j in range(2, len(df.index)+2):
            ws[f'{chr(ord("A") + i)}{j}'].style = wb_style_cells

    rd = ws.row_dimensions[1]
    rd.height = 22.32

    # Format length width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length
        ws.column_dimensions[column].width = adjusted_width

    wb.save(path_to_file)
    wb.close()


if __name__ == "__main__":
    path_to_file = sys.argv[1].replace('\\','/')
    
    print('Reading excel...', end='')
    df = read_excel(path_to_file)
    print('Done')

    print('Calculating...', end='')
    df_sum_temp, df_avg_temp = cal_report(df)
    print('Done')

    print('Writing to excel...', end='')
    df = rearrange_df(df_sum_temp, df_avg_temp)
    write_to_excel(df, path_to_file)
    format_excel(path_to_file, df)
    print('Done')

    print('All done')

    del(rows)
    del(sum_rows)
    del(avg_rows)
